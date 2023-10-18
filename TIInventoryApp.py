import tkinter as tk
from tkinter import ttk
import openpyxl
from tkcalendar import DateEntry
from datetime import datetime

setor_colaboradores = {
  "Almoxarifado": ["Jose Elson"],
  "Comercial":
  ["Eduardo Pancini", "Tamires Axis", "André Moreira", "Emanuelle Polonine"],
  "Caldeiraria": ["Estagiário", "Thiago Germano", "Marcelo Alves"],
  "Compras": ["Willian Benincá", "Giovani Cipriano"],
  "Contabilidade": ["Rosemberg", "Uederson", "Sandra"],
  "Controle de qualidade": ["Leandro Cereza", "Sala Kaizen-Lean"],
  "Diretoria": ["Izaias", "Leandro Cereza"],
  "Engenharia": [
    "Estagiário", "Rogélio Buson", "Estagiário B", "Wanderson Fortunato",
    "Edvaldo Caetano"
  ],
  "Financeiro": ["Jose Leal", "Walleska"],
  "PPCP": [
    "Romário Freitas", "José Geraldo Mota", "Victor Carvalho",
    "Willian Tiburcio", "Ariele Grechi", "Estagiário(Ariele)",
    "Estagiário(Zuqui)", "Leonardo Zuqui", "Estagiário(Canto)"
  ],
  "Recepção": ["Recepcionista"],
  "RG/SGI/SEG": ["Jorge Tavares", "Janine", "Jaquenily Morelli"],
  "TI": ["José Adilson", "Estagiário"]
  # ... e assim por diante para os outros setores
}


class TIInventoryApp:

  def __init__(self, root):
    self.root = root
    self.root.title("ITAMIL ITAPEMIRIM INDÚSTRIA LTDA TI")

    self.style = ttk.Style()
    self.style.theme_use("clam")

    # Criar e configurar a tabela de exibição
    self.tree = ttk.Treeview(root,
                             columns=("Setor", "Colaborador", "Equipamento",
                                      "Retirada", "Devolução"),
                             show="headings")
    self.tree.heading("Equipamento", text="Equipamento")
    self.tree.heading("Setor", text="Setor")
    self.tree.heading("Colaborador", text="Colaborador")
    self.tree.heading("Retirada", text="Retirada")
    self.tree.heading("Devolução", text="Devolução")
    self.tree.grid(row=0,
                   column=0,
                   columnspan=2,
                   padx=10,
                   pady=10,
                   sticky="nsew")
    #ESTILOS
    self.style.configure("TLabel", font=("Helvetica", 8))
    self.style.configure("TButton",
                         font=("Helvetica", 8),
                         background="#4CAF50",
                         foreground="white")
    self.style.configure("TEntry", font=("Helvetica", 8))

    # Botões e entradas

    # Seletor de Setores usando ttk.Combobox
    self.setores = [
      "Almoxarifado", "Comercial", "Caldeiraria", "Compras", "Contabilidade",
      "Controle de qualidade", "Diretoria", "Engenharia", "Financeiro", "PPCP",
      "Procedimentos", "Recepção", "RG/SGI/SEG", "TI"
    ]
    self.setor_combobox = ttk.Combobox(root,
                                       values=self.setores,
                                       state="readonly")
    self.setor_combobox.grid(row=1, column=1, padx=10, pady=5, sticky="w")
    self.setor_combobox.set("Selecione um setor")
    self.setor_combobox.bind("<<ComboboxSelected>>", self.update_colaboradores)

    # self.setor_label = tk.Label(root, text="Setor:", font=("Helvetica", 12))
    # self.setor_label.grid(row=2, column=0, padx=10, pady=5, sticky="e")
    #self.setor_entry = tk.Entry(root, font=("Helvetica",10))
    # self.setor_entry.grid(row=2, column=1, padx=10, pady=5, sticky="w")

    self.colaborador_combobox = ttk.Combobox(root, state="readonly")
    self.colaborador_combobox.set("Selecione um colaborador")
    self.colaborador_combobox.grid(row=2,
                                   column=1,
                                   padx=10,
                                   pady=5,
                                   sticky="w")

    self.item_label = tk.Label(root,
                               text="Equipamento:",
                               font=("Helvetica", 10))
    self.item_label.grid(row=3, column=0, padx=10, pady=5, sticky="e")
    self.item_entry = tk.Entry(root, font=("Helvetica", 10))
    self.item_entry.grid(row=3, column=1, padx=10, pady=5, sticky="w")

    self.retirada_label = tk.Label(root,
                                   text="Retirada:",
                                   font=("Helvetica", 10))
    self.retirada_label.grid(row=4, column=0, padx=10, pady=5, sticky="e")
    self.retirada_entry = DateEntry(root, date_pattern="dd/mm/yyyy")
    self.retirada_entry.grid(row=4, column=1, padx=10, pady=5, sticky="w")

    self.data_devolucao_label = tk.Label(root,
                                         text="Devolução:",
                                         font=("Helvetica", 10))
    self.data_devolucao_label.grid(row=5,
                                   column=0,
                                   padx=10,
                                   pady=5,
                                   sticky="e")
    self.data_devolucao_entry = DateEntry(root, date_pattern="dd/mm/yyyy")
    self.data_devolucao_entry.grid(row=5,
                                   column=1,
                                   padx=10,
                                   pady=5,
                                   sticky="w")

    self.add_button = tk.Button(root,
                                text="Adicionar",
                                command=self.add_item,
                                font=("Helvetica", 12),
                                bg="#4CAF50",
                                fg="white")
    self.add_button.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

    self.remove_button = tk.Button(root,
                                   text="Remover",
                                   command=self.remove_item,
                                   font=("Helvetica", 12),
                                   bg="#af4c4c",
                                   fg="white")
    self.remove_button.grid(row=7, column=0, columnspan=2, padx=10, pady=10)

    # Carregar dados do Excel
    self.load_data()

  def update_colaboradores(self, event):
    setor = self.setor_combobox.get()
    colaboradores = setor_colaboradores.get(setor, [])
    self.colaborador_combobox["values"] = colaboradores
    self.colaborador_combobox.set("")

  def add_item(self):
    item = self.item_entry.get()
    setor = self.setor_combobox.get()
    colaborador = self.colaborador_combobox.get()
    retirada = self.retirada_entry.get()
    data_devolucao = self.data_devolucao_entry.get()

    if item and setor and colaborador and retirada:
      self.tree.insert("",
                       "end",
                       values=(item, setor, colaborador, retirada,
                               data_devolucao))

      # Adicionar dados ao arquivo Excel
      self.add_to_excel(item, setor, colaborador, retirada, data_devolucao)
    colaboradores = setor_colaboradores.get(setor, [])
    if colaboradores:
      self.colaborador_combobox["values"] = colaboradores
      self.colaborador_combobox.set("")
    else:
      self.colaborador_combobox["values"] = []
      # Limpar as entradas

  # Limpar as entradas
    self.item_entry.delete(0, "end")
    self.setor_combobox.delete(0, "end")
    self.colaborador_combobox.delete(0, "end")
    self.retirada_entry.delete(0, tk.END)
    self.data_devolucao_entry.delete(0, tk.END)

  def remove_item(self):
    selected_items = self.tree.selection()
    for item_id in selected_items:
      self.tree.delete(item_id)
      # Remover dados do arquivo Excel
      self.remove_from_excel(item_id)

  def load_data(self):
    try:
      wb = openpyxl.load_workbook("inventory.xlsx")
      sheet = wb.active
      for row in sheet.iter_rows(min_row=2, values_only=True):
        self.tree.insert("", "end", values=row)

        item, setor, colaborador, retirada, data_devolucao = row
        if setor not in setor_colaboradores:
          setor_colaboradores[setor] = []
        if colaborador not in setor_colaboradores[setor]:
          setor_colaboradores[setor].append(colaborador)

      wb.close()

    except FileNotFoundError:
      pass

  def add_to_excel(self, item, setor, colaborador, retirada, data_devolucao):
    try:
      wb = openpyxl.load_workbook("inventory.xlsx")
      sheet = wb.active
      sheet.append([item, setor, colaborador, retirada, data_devolucao])
      wb.save("inventory.xlsx")
      wb.close()

      if setor not in setor_colaboradores:
        setor_colaboradores[setor] = []
      if colaborador not in setor_colaboradores[setor]:
        setor_colaboradores[setor].append(colaborador)
    except FileNotFoundError:
      wb = openpyxl.Workbook()
      sheet = wb.active
      sheet.append(["Item", "Setor", "Colaborador", "Retirada", "Devolução"])
      sheet.append([item, setor, colaborador, retirada, data_devolucao])
      wb.save("inventory.xlsx")
      wb.close()

  def remove_from_excel(self, item_id):
    try:
      wb = openpyxl.load_workbook("inventory.xlsx")
      sheet = wb.active
      row_index = int(item_id.split("I")[1]) - 1  # Ajuste de índice base 0
      sheet.delete_rows(row_index +
                        2)  # +2 para considerar cabeçalho e índice base 1
      wb.save("inventory.xlsx")
      wb.close()
    except FileNotFoundError:
      pass

  def clear_entries(self):
    self.item_entry.delete(0, "end")
    self.setor_entry.delete(0, "end")
    self.colaborador_entry.delete(0, "end")
    self.retirada_entry.delete(0, "end")
    self.data_devolucao_entry.delete(0, "end")


if __name__ == "__main__":
  root = tk.Tk()
  app = TIInventoryApp(root)
  root.mainloop()
